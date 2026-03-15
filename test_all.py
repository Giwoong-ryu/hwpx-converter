"""전체 내부 테스트"""
import tempfile, os, json, csv, sys
from pathlib import Path

errors = []
passed = 0

from core.xml_utils import reset_id_counter
from core.json_to_section import generate_section_xml, create_header, _TEMPLATE_STYLES
from core.build_hwpx import build, AVAILABLE_TEMPLATES
from excel_parser import parse_file
from form_filler import analyze_template, fill_template, parse_values_file
from template_analyzer import analyze_hwpx, extract_header, extract_section
from page_guard import collect_metrics, compare_metrics
import openpyxl
import re

# 공통 테스트 문서
DOC = {
    "document": {"title": "test", "page_width_hu": 42520},
    "sections": [
        {"type": "paragraph", "content": "Hello", "align": "CENTER", "style": {"bold": True, "font_size_pt": 16}},
        {"type": "table", "table": {"rows": 2, "cols": 2, "col_widths_ratio": [50, 50],
            "cells": [
                {"row":0,"col":0,"text":"A","style":{"bold":True,"align":"CENTER","is_header":True}},
                {"row":0,"col":1,"text":"B","style":{"bold":True,"align":"CENTER","is_header":True}},
                {"row":1,"col":0,"text":"1","style":{"align":"LEFT"}},
                {"row":1,"col":1,"text":"2","style":{"align":"LEFT"}},
            ]}}
    ]
}

def run(name, fn):
    global passed
    try:
        fn()
        print(f"[OK] {name}")
        passed += 1
    except Exception as e:
        errors.append(f"{name}: {e}")
        print(f"[FAIL] {name}: {e}")


def test_imports():
    from ocr_engine import ocr_with_gemini, ocr_with_easyocr
run("1. imports", test_imports)


def test_build_templates():
    for t in ["auto", "report", "gonmun", "minutes", "proposal"]:
        reset_id_counter()
        with tempfile.TemporaryDirectory() as d:
            xml_dir = os.path.join(d, "xml"); os.makedirs(xml_dir)
            generate_section_xml(DOC, xml_dir)
            create_header(xml_dir, template=t)
            out = os.path.join(d, f"{t}.hwpx")
            errs = build(template=t, header_override=Path(xml_dir)/"header.xml",
                        section_override=Path(xml_dir)/"section0.xml",
                        title="test", creator="test", output=Path(out))
            assert not errs, f"build {t}: {errs}"
            assert os.path.getsize(out) > 5000
run("2. build 5 templates", test_build_templates)


def test_excel_parser():
    wb = openpyxl.Workbook(); ws = wb.active
    ws["A1"] = "Name"; ws["B1"] = "Value"
    ws["A2"] = "Test"; ws["B2"] = "123"
    p = tempfile.mktemp(suffix=".xlsx"); wb.save(p); wb.close()
    r = parse_file(p)
    assert "sections" in r and len(r["sections"]) > 0
run("3. Excel parser", test_excel_parser)


def test_csv_parser():
    p = tempfile.mktemp(suffix=".csv")
    with open(p, "w", encoding="utf-8-sig", newline="") as f:
        csv.writer(f).writerows([["Col1","Col2"],["A","B"]])
    r = parse_file(p)
    assert "sections" in r
run("4. CSV parser", test_csv_parser)


def test_form_analyze_fill():
    reset_id_counter()
    form_doc = {
        "document": {"title": "form", "page_width_hu": 42520},
        "sections": [{"type": "table", "table": {"rows": 2, "cols": 2, "col_widths_ratio": [40,60],
            "cells": [
                {"row":0,"col":0,"text":"Name","style":{"bold":True,"align":"CENTER","is_header":True}},
                {"row":0,"col":1,"text":"{{name}}","style":{"align":"LEFT"}},
                {"row":1,"col":0,"text":"Dept","style":{"bold":True,"align":"CENTER","is_header":True}},
                {"row":1,"col":1,"text":"{{dept}}","style":{"align":"LEFT"}},
            ]}}]
    }
    with tempfile.TemporaryDirectory() as d:
        xml_dir = os.path.join(d, "xml"); os.makedirs(xml_dir)
        generate_section_xml(form_doc, xml_dir)
        create_header(xml_dir, template="auto")
        tpl = os.path.join(d, "form.hwpx")
        build(template="auto", header_override=Path(xml_dir)/"header.xml",
              section_override=Path(xml_dir)/"section0.xml",
              title="form", creator="test", output=Path(tpl))

        fields, _ = analyze_template(tpl)
        assert set(fields) == {"name", "dept"}, f"Got: {fields}"

        out_path, log = fill_template(tpl, {"name": "John", "dept": "Dev"})
        assert os.path.exists(out_path)
        assert "page_guard" in log

        remaining, _ = analyze_template(out_path)
        assert len(remaining) == 0
run("5. form analyze + fill + page_guard", test_form_analyze_fill)


def test_analyze_hwpx():
    reset_id_counter()
    with tempfile.TemporaryDirectory() as d:
        xml_dir = os.path.join(d, "xml"); os.makedirs(xml_dir)
        generate_section_xml(DOC, xml_dir)
        create_header(xml_dir, template="report")
        tpl = os.path.join(d, "t.hwpx")
        build(template="report", header_override=Path(xml_dir)/"header.xml",
              section_override=Path(xml_dir)/"section0.xml",
              title="t", creator="t", output=Path(tpl))
        report = analyze_hwpx(tpl)
        assert "borderFill" in report and "charPr" in report
run("6. analyze_hwpx", test_analyze_hwpx)


def test_extract_header_section():
    reset_id_counter()
    with tempfile.TemporaryDirectory() as d:
        xml_dir = os.path.join(d, "xml"); os.makedirs(xml_dir)
        generate_section_xml(DOC, xml_dir)
        create_header(xml_dir, template="auto")
        tpl = os.path.join(d, "t.hwpx")
        build(template="auto", header_override=Path(xml_dir)/"header.xml",
              section_override=Path(xml_dir)/"section0.xml",
              title="t", creator="t", output=Path(tpl))
        hdr = extract_header(tpl, d)
        sec = extract_section(tpl, d)
        assert os.path.exists(hdr) and os.path.getsize(hdr) > 1000
        assert os.path.exists(sec) and os.path.getsize(sec) > 500
run("7. extract_header/section", test_extract_header_section)


def test_values_excel():
    wb = openpyxl.Workbook(); ws = wb.active
    ws["A1"] = "name"; ws["B1"] = "Alice"
    ws["A2"] = "dept"; ws["B2"] = "HR"
    p = tempfile.mktemp(suffix=".xlsx"); wb.save(p); wb.close()
    assert parse_values_file(p) == {"name": "Alice", "dept": "HR"}
run("8. values Excel", test_values_excel)


def test_values_csv():
    p = tempfile.mktemp(suffix=".csv")
    with open(p, "w", encoding="utf-8-sig", newline="") as f:
        csv.writer(f).writerows([["title","Report"],["date","2026-03-15"]])
    assert parse_values_file(p) == {"title": "Report", "date": "2026-03-15"}
run("9. values CSV", test_values_csv)


def test_values_json():
    p = tempfile.mktemp(suffix=".json")
    with open(p, "w", encoding="utf-8") as f:
        json.dump({"k1": "v1", "k2": "v2"}, f)
    assert parse_values_file(p) == {"k1": "v1", "k2": "v2"}
run("10. values JSON", test_values_json)


def test_values_txt():
    p = tempfile.mktemp(suffix=".txt")
    with open(p, "w", encoding="utf-8") as f:
        f.write("name: Bob\nage: 30\n")
    assert parse_values_file(p) == {"name": "Bob", "age": "30"}
run("11. values TXT", test_values_txt)


def test_values_docx():
    from docx import Document
    doc = Document()
    t = doc.add_table(rows=2, cols=2)
    t.cell(0,0).text = "Co"; t.cell(0,1).text = "Acme"
    t.cell(1,0).text = "CC"; t.cell(1,1).text = "KR"
    p = tempfile.mktemp(suffix=".docx"); doc.save(p)
    assert parse_values_file(p) == {"Co": "Acme", "CC": "KR"}
run("12. values DOCX", test_values_docx)


def test_page_guard_self():
    reset_id_counter()
    with tempfile.TemporaryDirectory() as d:
        xml_dir = os.path.join(d, "xml"); os.makedirs(xml_dir)
        generate_section_xml(DOC, xml_dir)
        create_header(xml_dir, template="report")
        f1 = os.path.join(d, "a.hwpx")
        build(template="report", header_override=Path(xml_dir)/"header.xml",
              section_override=Path(xml_dir)/"section0.xml",
              title="a", creator="t", output=Path(f1))
        m = collect_metrics(Path(f1))
        errs = compare_metrics(m, m, 0.15, 0.25)
        assert len(errs) == 0, f"Self-compare: {errs}"
run("13. page_guard self-compare", test_page_guard_self)


def test_template_no_dupes():
    for t_name in _TEMPLATE_STYLES:
        reset_id_counter()
        with tempfile.TemporaryDirectory() as d:
            xml_dir = os.path.join(d, "xml"); os.makedirs(xml_dir)
            generate_section_xml(DOC, xml_dir)
            create_header(xml_dir, template=t_name)
            content = open(os.path.join(xml_dir, "header.xml"), "r", encoding="utf-8").read()
            count = len(re.findall(r'borderFill id="7"', content))
            assert count == 1, f"{t_name}: bf7 x{count}"
run("14. template styles no dupes", test_template_no_dupes)


# RESULT
print()
print("=" * 50)
if errors:
    print(f"RESULT: {passed} passed, {len(errors)} failed")
    for e in errors:
        print(f"  [FAIL] {e}")
    sys.exit(1)
else:
    print(f"ALL {passed} TESTS PASSED")
