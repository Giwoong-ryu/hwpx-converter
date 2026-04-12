"""
AI 매핑 모듈 - 양식 필드를 사용자 자료 기반으로 자동 채움 (통합 파이프라인)

필드 타입별 처리:
  [EXACT] 정밀 필드 (숫자/날짜/이름/코드) - 소스에서 정확 복사, 변형 금지
  [GEN]   서술 필드 (설명/소개/계획 등)   - 소스 기반으로 맥락에 맞게 작성
"""
import io
import json
import os
import re
import sys

# Windows cp949 콘솔에서 특수 Unicode 문자 출력 시 발생하는 UnicodeEncodeError 방지
# FastAPI 서버 로그 및 테스트 환경 모두 적용
if sys.platform == "win32" and hasattr(sys.stdout, "buffer"):
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
    except Exception:
        pass

from google import genai
from google.genai import types

MODEL_NAME = os.environ.get("DOCFLOW_MODEL", "gemini-2.5-flash")

# ── 필드 타입 자동 분류 ──
_EXACT_PATTERNS = re.compile(
    r"금액|합계|총액|단가|비용|가격|원가|수수료|급여|수당|세금|부가세|공급가"
    r"|일자|날짜|일시|기간|발행일|작성일|시작일|종료일|마감|납기"
    r"|전화|연락|핸드폰|휴대|팩스|FAX|H\.P|HP"
    r"|사업자|등록번호|주민등록|법인번호"
    r"|성명|이름|대표자|담당자"
    r"|주소|소재지|우편번호|계좌|은행|예금주"
    r"|수량|단위|규격|품명|품목"
    r"|E-MAIL|EMAIL|이메일", re.IGNORECASE)

_GEN_PATTERNS = re.compile(
    r"내용|설명|비고|특기|소개|목적|사유|의견|요약|개요"
    r"|지원동기|자기소개|경력기술|직무|역할|성과|포부"
    r"|비전|전략|계획|방안|제안|결론"
    r"|성장과정|장단점|입사", re.IGNORECASE)


def _classify_field_type(label: str) -> str:
    """양식 라벨을 분석하여 필드 타입을 반환한다.
    Returns: "EXACT" / "GEN" / "" (분류 불가)
    """
    if not label:
        return ""
    if _EXACT_PATTERNS.search(label):
        return "EXACT"
    if _GEN_PATTERNS.search(label):
        return "GEN"
    return ""


def _calc_max_chars(width_hwpunit: int, height_hwpunit: int, font_sz_hwpunit: int = 1000) -> int:
    """셀 크기(HWPUNIT) + 폰트 크기로부터 최대 허용 글자 수를 추정한다.

    HWPUNIT 규칙:
    - 1pt = 100 HWPUNIT, 1mm ≈ 283 HWPUNIT
    - 한글 1글자 폭 ≈ 폰트 크기(정사각형)
    - 줄 높이 ≈ 폰트 크기 × 1.6

    여유율 0.85로 안전 마진 확보. 0이 반환되면 제약 없음으로 해석.
    """
    if width_hwpunit <= 0 or height_hwpunit <= 0 or font_sz_hwpunit <= 0:
        return 0
    # 영문은 한글의 약 0.5폭이지만 한글 기준으로 보수적 계산
    char_w = max(font_sz_hwpunit, 500)  # 최소 5pt 방어
    line_h = char_w * 1.6
    chars_per_line = max(int(width_hwpunit / char_w), 1)
    lines = max(int(height_hwpunit / line_h), 1)
    return int(chars_per_line * lines * 0.85)


def _strip_field_tags(key: str) -> str:
    """AI 응답 키에서 [EXACT], [GEN], [max=N] 태그를 제거한다."""
    # 복합 태그 [EXACT,max=420] 또는 단일 태그 [GEN]/[max=100] 모두 처리
    return re.sub(r"\[[A-Za-z0-9,=]+\]", "", key).strip()


SYSTEM_PROMPT = """\
당신은 한글 문서 양식 작성 도우미입니다.
사용자가 양식(HWP/HWPX)의 테이블 구조와 참고 자료를 제공하면,
자료를 이해하여 양식의 값 필드를 채웁니다.

양식은 테이블 구조로 제공됩니다. 각 행에서:
- [H] 표시가 있는 셀 = 라벨/헤더 (절대 교체 금지)
- __N 접미사가 붙은 셀 = 같은 유형의 N번째 항목 (예: 회사명__1, 회사명__2)
- [EXACT] 태그가 있는 셀 = 정밀 필드 (소스에서 정확히 복사, 변형/추측 금지)
- [GEN] 태그가 있는 셀 = 서술 필드 (소스를 바탕으로 자연스럽게 작성)
- [max=N] 태그 = 해당 셀의 최대 허용 글자 수. 공백 포함 N자 이내 필수.
- 태그는 복합될 수 있음: [GEN,max=420], [EXACT,max=30]
- 태그 없는 일반 셀 = 값 (교체 대상)

규칙:
1. 라벨/헤더 셀([H])은 절대 교체하지 마세요.
2. [EXACT] 필드: 소스 자료에서 정확히 찾아 복사하세요. 숫자, 날짜, 코드, 이름 등은
   한 글자도 변형하지 마세요. 소스에 없으면 빈 문자열("")을 반환하세요.
3. [GEN] 필드: 소스 자료를 이해하고 양식 맥락에 맞게 작성하세요.
   소스에 없는 내용을 추가하거나 날조하지 마세요.
4. __N 접미사 셀은 소스에 실제로 있는 N번째 항목만 채우세요.
   - 소스에서 해당 항목의 개수를 먼저 세세요. N개라면 __1~__N만 생성합니다.
   - 소스에 없는 N번째 항목은 JSON에 포함하지 마세요.
   - 절대 같은 항목을 __2, __3에 복사하지 마세요.
5. 경력, 학력처럼 시기/날짜가 있는 항목은 과거(오래된 순)에서 현재 순으로 위(__1)부터 채우세요.
6. 반드시 JSON만 반환하세요. 설명이나 마크다운 없이.
7. __N 접미사가 있는 셀은 JSON 키에도 반드시 __N을 포함하세요.
8. 매핑 가능한 필드는 최대한 빠짐없이 채우세요.
9. 섹션명__1 형태 셀은 자기소개서 등 섹션형 양식의 본문 슬롯입니다.
   - 해당 섹션의 본문 내용을 소스 자료를 바탕으로 작성하세요.
   - 내용이 길더라도 줄바꿈 없이 한 문자열로 작성하세요.
   - [max=N]이 붙어 있으면 그 글자 수 이내로 핵심만 담아 작성하세요.
   - 섹션당 __1만 존재합니다. __2, __3은 만들지 마세요.
10. "제목 :__N", "X :__N" 형태 셀은 append 라벨 슬롯입니다.
    - 양식의 "X : ___" 패턴에서 콜론 뒤 공백에 들어갈 값입니다.
    - JSON 키를 그대로 써야 하며 값에는 "X :" 라벨을 포함하지 마세요.
      (예: 키 "제목 :__1", 값 "현장 실무와 자동화 역량의 결합")
    - 섹션마다 __1, __2, __3, __4로 번호가 붙으므로 소스 자료와 섹션명을
      매칭해서 각기 다른 값을 생성하세요. 같은 값 복사 금지.
    - 자기소개서의 경우 __1~__4는 순서대로 성장과정/성격/지원동기/포부 소제목입니다.
    - [max=N]이 붙으므로 반드시 그 한도 내 짧고 인상적인 한 줄 제목으로.
11. [max=N] 글자 수 제한 — 양식은 한 페이지에 담도록 설계되어 있습니다.
    - N자 초과 절대 금지. 공백과 문장부호 모두 포함해서 N자 이내로 작성.
    - 서술형([GEN]) 필드: N자 한도 내에서 핵심만 압축. 긴 문장보다 짧고 밀도 있게.
    - 정밀([EXACT]) 필드: 소스 값이 N자를 초과하면 빈 문자열 반환 (임의 절단 금지).
12. JSON 키 형식 — 반드시 "라벨명" 또는 "라벨명__N" 형식만 허용합니다.
    - "공급받는자용_라벨명", "공급자용_라벨명", "섹션명_라벨명" 같은 섹션 접두사 절대 금지.
    - 양식에 같은 라벨이 2번 나오면 "라벨명__1", "라벨명__2"로 구분하세요.
    - 두 벌 양식(공급받는자용/공급자용)에서 같은 라벨은 __1(왼쪽), __2(오른쪽)로 구분합니다.
13. 키는 반드시 양식 [H] 라벨 텍스트 그대로입니다. 라벨 의미를 해석해서 다른 이름을 만들지 마세요.
    - 예: [H]합  계 → 키='합  계' (O), '총 견적 금액'(X), '공급가액합계'(X)
    - 예: [H]귀하 앞의 빈 셀 → 수신인 이름을 채우되 키='귀하__N' (O), '고객명'(X), '수신자'(X)
    - 귀하/수신자/귀중 앞 빈 셀은 수신인 이름 슬롯입니다. 반드시 채우세요.
14. 학년 필드는 반드시 단축형으로 입력하세요 (초등학교 제외).
    - 중학교 1학년 → '중1', 중학교 2학년 → '중2', 중학교 3학년 → '중3'
    - 고등학교 1학년 → '고1', 고등학교 2학년 → '고2', 고등학교 3학년 → '고3'
    - '중학교 2학년', '2학년' 같은 풀네임 형식은 절대 금지."""

USER_PROMPT = """\
[양식 구조]
{fields}

[사용자 제공 자료]
{content}

위 양식의 빈 셀을 사용자 자료를 바탕으로 채워 JSON으로 반환하세요.
형식: {{"필드명": "채울 내용", ...}}

핵심 원칙:
- [EXACT] 필드는 소스에서 정확히 복사. 없으면 빈 문자열.
- [GEN] 필드는 소스를 이해하여 양식에 맞게 작성. 소스 밖 내용 금지.
- [max=N] 태그: 공백 포함 N자 이내 필수. 초과 시 양식이 망가집니다.
- [H] 라벨 텍스트는 절대 JSON 키로 쓰지 마세요.

매핑 전략:
0. [반복 항목 선추출 — 필수] 품목/경력/학력처럼 여러 행이 있는 항목은
   JSON 맨 앞에 "_목록" 키로 전체 목록을 먼저 나열한 뒤 슬롯을 채우세요.
   이 키는 실제 양식에 삽입되지 않으므로 빠짐없이 정직하게 나열하세요.
   (견적서의 품목1~N, 이력서의 경력1~N 등 모두 해당. 목록 수만큼 __N 슬롯 생성.)
1. [H] 라벨을 JSON 키로 사용해 해당 빈 칸에 넣을 값을 지정하세요.
2. __N 접미사로 여러 항목 구분. 0번 단계에서 나열한 목록 수만큼만 생성.
3. 셀에 이미 텍스트가 있는 경우 그 텍스트를 키로 쓰세요.
4. 문맥상 명확하면 추론하여 채우세요.
5. 매핑 가능한 모든 필드를 최대한 빠짐없이 채우세요.
6. 합  계/합계 셀에는 부가세 포함 최종 총액을 넣으세요. 소계(공급가액합계, VAT 제외)가 아닙니다."""


def _format_structured_fields(structured, extra_labels=None):
    """구조화된 필드 데이터를 AI 프롬프트용 텍스트로 변환한다.

    테이블은 마크다운 테이블 형태, bold/bg 셀은 [H] 태그로 표시.
    중복 값 셀에는 __N 인덱스를 붙여 AI가 순서대로 개별 매핑하도록 한다.

    Args:
        structured: extract_structured_fields 결과
        extra_labels: 일반 텍스트로 되어있지만 [H]로 취급할 추가 라벨 집합
            (invoice_style 양식에서 InvoiceProcessor의 INVOICE_LABELS를 넘긴다)
    """
    _SKIP = {"□", "☑", "※", "①", "②", "③", "④", "⑤", "⑥", "⑦", "⑧", "⑨", "⑩", "☐", "○", "●"}

    # 서식 내 소라벨 (bold/bg 없어도 헤더로 처리)
    _SUB_LABELS = {
        "H.P", "HP", "E-MAIL", "E.MAIL", "EMAIL", "TEL", "FAX",
        "전화", "휴대폰", "이메일", "홈페이지",
        "상", "중", "하", "상/중/하",
    }

    # extra_labels (invoice 라벨)도 소라벨로 취급 - 정규화해서 매칭용 캐시 생성
    extra_labels_norm: set[str] = set()
    if extra_labels:
        extra_labels_norm = {re.sub(r"\s+", "", lbl).lower() for lbl in extra_labels}

    def _is_extra_label(text: str) -> bool:
        if not extra_labels_norm:
            return False
        t_norm = re.sub(r"\s+", "", text).lower()
        return t_norm in extra_labels_norm

    # 1패스: 값 셀 텍스트 빈도 계산 (중복 여부 파악)
    text_freq: dict[str, int] = {}
    for table in structured["tables"]:
        for row in table["rows"]:
            for cell in row:
                t = cell["text"].strip()
                if (t and t not in _SKIP and t not in _SUB_LABELS
                        and not cell["bold"] and not cell["bg"]
                        and not _is_extra_label(t)):
                    text_freq[t] = text_freq.get(t, 0) + 1

    # "X:" 패턴 (예: "제목 :") 라벨 append 슬롯 카운터 (섹션마다 증가)
    _APPEND_PAT = re.compile(r'[:：]\s*$')
    label_append_seen: dict[str, int] = {}

    # 2패스: 마크다운 테이블 생성 (중복 셀만 __N 추가)
    lines = []
    text_seen: dict[str, int] = {}

    for ti, table in enumerate(structured["tables"]):
        lines.append(f"[표{ti+1}]")

        # 1컬럼 섹션형 테이블(자기소개서 등) 탐지: 모든 행이 1셀
        is_single_col = all(len(row) == 1 for row in table["rows"])
        current_section: str | None = None
        section_empty_count: dict[str, int] = {}
        seen_sublabel: bool = False  # 섹션 헤더 뒤 bg 라벨(제목 :)을 봤는지

        last_header_label = ""  # 직전 헤더 라벨 (필드 타입 분류용)

        for row in table["rows"]:
            cells = []
            for cell in row:
                text = cell["text"].strip()
                if not text or text in _SKIP:
                    cells.append("")
                    continue
                if cell["bold"] or cell["bg"] or text in _SUB_LABELS or _is_extra_label(text):
                    # "X:" 패턴 append 라벨: 기존 텍스트 뒤에 값을 이어붙이는 슬롯
                    # 예: "제목 : " → 섹션별 소제목이 콜론 뒤로 들어감
                    is_append_label = (
                        cell["bg"] and not cell["bold"]
                        and _APPEND_PAT.search(text)
                        and len(text) <= 20
                    )
                    if is_append_label:
                        cell_cap = _calc_max_chars(
                            cell.get("width", 0),
                            cell.get("height", 0),
                            cell.get("font_sz", 1000),
                        )
                        free_chars = max(cell_cap - len(text), 0)
                        if free_chars >= 5:  # 5자 이상 여유 있어야 append 슬롯으로 승격
                            label_append_seen[text] = label_append_seen.get(text, 0) + 1
                            idx = label_append_seen[text]
                            tag = f"[GEN,max={free_chars}]"
                            cells.append(f"{tag}{text}__{idx}")
                            # 섹션형 로직에 본문 슬롯 생성 신호
                            if is_single_col:
                                seen_sublabel = True
                            continue  # last_header_label 갱신 스킵 (섹션명 유지)

                    cells.append(f"[H]{text}")
                    last_header_label = text  # 헤더 라벨 추적
                    if is_single_col:
                        if cell["bold"]:
                            current_section = text
                            section_empty_count[text] = 0
                            seen_sublabel = False
                        elif cell["bg"] and text:
                            seen_sublabel = True
                else:
                    # 인접 헤더 라벨 기반 필드 타입 태그 + 셀 크기 기반 max_chars
                    field_tag = _classify_field_type(last_header_label)
                    max_chars = _calc_max_chars(
                        cell.get("width", 0),
                        cell.get("height", 0),
                        cell.get("font_sz", 1000),
                    )
                    tag_parts = []
                    if field_tag:
                        tag_parts.append(field_tag)
                    if max_chars > 0:
                        tag_parts.append(f"max={max_chars}")
                    tag_prefix = f"[{','.join(tag_parts)}]" if tag_parts else ""

                    if text_freq.get(text, 1) > 1:
                        text_seen[text] = text_seen.get(text, 0) + 1
                        cells.append(f"{tag_prefix}{text}__{text_seen[text]}")
                    else:
                        cells.append(f"{tag_prefix}{text}")

            # 1컬럼 섹션형: 섹션 헤더 뒤의 "본문 슬롯"을 하나만 매핑한다.
            # 양식 구조: [헤더] → ["제목 :" 라벨] → [큰 본문 슬롯] → [작은 여백]
            # 여백 셀(h가 작은 빈 셀)은 매핑 대상에서 제외해야 넘침 방지.
            if (is_single_col
                    and current_section
                    and seen_sublabel
                    and len(cells) == 1
                    and not cells[0]
                    and not row[0].get("bg", False)):  # bg=True 빈 행은 구분자 → 스킵
                src_cell = row[0]
                h = src_cell.get("height", 0)
                # 본문 슬롯으로 인정할 최소 높이 임계값
                # - 한글 10pt 기준 1줄 ≈ 1600 HWPUNIT, 2줄 이상이면 본문으로 판단
                MIN_BODY_HEIGHT = 3000
                already_has_body = section_empty_count.get(current_section, 0) >= 1

                if h >= MIN_BODY_HEIGHT and not already_has_body:
                    section_empty_count[current_section] = 1
                    sec_max = _calc_max_chars(
                        src_cell.get("width", 0),
                        h,
                        src_cell.get("font_sz", 1000),
                    )
                    sec_tag_parts = ["GEN"]
                    if sec_max > 0:
                        sec_tag_parts.append(f"max={sec_max}")
                    sec_tag = f"[{','.join(sec_tag_parts)}]"
                    # __1 = 본문 슬롯 (form.py __\d+$ 정규식 호환)
                    cells = [f"{sec_tag}{current_section}__1"]
                # else: 여백/구분 셀 → cells=[""] 유지 → 매핑 안 됨

            if any(c for c in cells):
                lines.append("| " + " | ".join(cells) + " |")

    if structured.get("paragraphs"):
        lines.append("\n[본문]")
        for p in structured["paragraphs"]:
            p = p.strip()
            if p and p not in _SKIP:
                lines.append(f"- {p}")

    return "\n".join(lines)


def _call_with_retry(client, model_name, prompt, system_prompt, temperature,
                     response_schema=None, max_retries=2):
    """Gemini API 호출 + 429 재시도 + Structured Output"""
    config_kwargs = {
        "system_instruction": system_prompt,
        "temperature": temperature,
        "max_output_tokens": 32768,
    }
    if response_schema is not None:
        config_kwargs["response_mime_type"] = "application/json"
        config_kwargs["response_schema"] = response_schema

    for attempt in range(max_retries):
        try:
            return client.models.generate_content(
                model=model_name,
                contents=prompt,
                config=types.GenerateContentConfig(**config_kwargs),
            )
        except Exception as e:
            if "429" in str(e) and attempt < max_retries - 1:
                import time
                print(f"[ai/map] 429 rate limit, 3초 대기 후 재시도")
                time.sleep(3)
            else:
                raise


def _call_cached_with_retry(client, model_name, cache_name, prompt, temperature,
                            response_schema=None, max_retries=2):
    """캐시 사용 Gemini API 호출 + 429 재시도 + Structured Output"""
    config_kwargs = {
        "cached_content": cache_name,
        "temperature": temperature,
        "max_output_tokens": 32768,
    }
    if response_schema is not None:
        config_kwargs["response_mime_type"] = "application/json"
        config_kwargs["response_schema"] = response_schema

    for attempt in range(max_retries):
        try:
            return client.models.generate_content(
                model=model_name,
                contents=prompt,
                config=types.GenerateContentConfig(**config_kwargs),
            )
        except Exception as e:
            if "429" in str(e) and attempt < max_retries - 1:
                import time
                print(f"[ai/map] 429 rate limit, 3초 대기 후 재시도")
                time.sleep(3)
            else:
                raise


def _truncate_smart(text: str, max_chars: int) -> str:
    """max_chars 이내로 자른다. 가능하면 문장 경계에서 자연스럽게 자른다."""
    if len(text) <= max_chars:
        return text
    truncated = text[:max_chars]
    # 마지막 문장 부호에서 자르기 (60% 이상 위치 필요)
    threshold = max_chars * 0.6
    for sep in ['. ', '! ', '? ', '。', '다. ', '요. ', '다.', '요.']:
        idx = truncated.rfind(sep)
        if idx >= threshold:
            return truncated[:idx + len(sep)].rstrip()
    # 단어 경계에서 자르기
    idx = truncated.rfind(' ')
    if idx >= threshold:
        return truncated[:idx].rstrip()
    return truncated.rstrip()


def _collect_results(parsed):
    """파싱된 JSON에서 유효한 key-value 쌍만 수집. null/빈값은 제외.
    [max=N] 태그가 있으면 초과분을 문장 경계에서 자른다 (방어선).
    """
    results = {}
    for k, v in parsed.items():
        if not isinstance(k, str) or not k.strip():
            continue
        raw_key = k.strip()
        # max 태그 파싱 (제거 전에 먼저 추출)
        max_match = re.search(r"max=(\d+)", raw_key)
        max_chars = int(max_match.group(1)) if max_match else 0

        k = _strip_field_tags(raw_key)  # [EXACT]/[GEN]/[max=N] 태그 제거
        if v is None:
            continue  # null 값 = 소스에 없는 정보 → 스킵
        if isinstance(v, (int, float)):
            v = str(v)
        if not isinstance(v, str) or not v.strip():
            continue
        v = v.strip()
        # max_chars 초과 시 절단 (AI가 프롬프트 규칙을 어긴 경우 방어)
        if max_chars > 0 and len(v) > max_chars:
            original_len = len(v)
            v = _truncate_smart(v, max_chars)
            print(f"[verify] max={max_chars} 초과({original_len}자), 절단: {k}")
        results[k] = v
    return results


def _verify_exact_fields(mappings: dict, source_text: str) -> None:
    """정밀 필드([EXACT] 패턴)의 값이 소스에 실제 존재하는지 검증한다.
    소스에 없는 값은 경고 로그를 남기고 제거한다 (할루시네이션 방지).
    """
    if not source_text:
        return
    source_normalized = re.sub(r"\s+", "", source_text.lower())

    to_remove = []
    for key, value in mappings.items():
        if not _EXACT_PATTERNS.search(key):
            continue
        value_str = str(value)
        value_digits = re.sub(r"[^0-9]", "", value_str)
        value_normalized = re.sub(r"\s+", "", value_str.lower())

        if len(value_digits) >= 4:
            # 숫자 기반 검증 (금액, 전화번호, 사업자번호 등)
            source_digits = re.sub(r"[^0-9]", "", source_text)
            if value_digits not in source_digits:
                try:
                    print(f"[verify] 정밀 필드 불일치, 제거: {key}={value_str}")
                except UnicodeEncodeError:
                    print(f"[verify] 정밀 필드 불일치, 제거: {key}=(인코딩 오류)")
                to_remove.append(key)
        elif len(value_normalized) >= 2:
            # 텍스트 기반 검증 (이름, 주소 등)
            if value_normalized not in source_normalized:
                try:
                    print(f"[verify] 정밀 필드 불일치, 제거: {key}={value_str}")
                except UnicodeEncodeError:
                    print(f"[verify] 정밀 필드 불일치, 제거: {key}=(인코딩 오류)")
                to_remove.append(key)

    for k in to_remove:
        del mappings[k]
    if to_remove:
        print(f"[verify] 정밀 필드 {len(to_remove)}개 제거됨")


def _get_api_key():
    """환경변수에서 Gemini API 키를 가져온다."""
    key = os.environ.get("GEMINI_API_KEY")
    if not key:
        # .env 파일 체크
        env_path = os.path.join(os.path.dirname(__file__), ".env")
        if os.path.exists(env_path):
            with open(env_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line.startswith("GEMINI_API_KEY="):
                        key = line.split("=", 1)[1].strip().strip("\"'")
                        break
    return key


def _parse_json_response(text, structured_output=False):
    """Gemini 응답에서 JSON을 추출한다. 잘린 JSON도 최대한 복구한다."""
    text = text.strip()

    # Structured Output 모드: 이미 유효한 JSON이므로 바로 파싱
    if structured_output:
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            pass  # 폴백: 기존 복구 로직

    # 마크다운 코드블록 제거
    text = re.sub(r"```json\s*", "", text)
    text = re.sub(r"```\s*", "", text)
    text = text.strip()

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # 첫 { 부터 마지막 } 까지 추출
    first = text.find("{")
    last = text.rfind("}")
    if first >= 0 and last > first:
        try:
            return json.loads(text[first:last + 1])
        except json.JSONDecodeError:
            pass

    # 잘린 JSON 복구 시도: 마지막 완전한 key-value 쌍까지 자르고 } 추가
    if first >= 0:
        fragment = text[first:]
        # 마지막 완전한 "value" 다음의 쉼표나 줄바꿈까지 찾기
        last_quote = fragment.rfind('"')
        if last_quote > 0:
            # 마지막 따옴표 이후 잘라내고 } 추가
            candidate = fragment[:last_quote + 1]
            # 마지막 항목 뒤 쉼표 제거
            candidate = candidate.rstrip().rstrip(",")
            candidate += "\n}"
            try:
                return json.loads(candidate)
            except json.JSONDecodeError:
                pass

    # 배열 시도
    first = text.find("[")
    last = text.rfind("]")
    if first >= 0 and last > first:
        try:
            return json.loads(text[first:last + 1])
        except json.JSONDecodeError:
            pass

    return None


def _extract_html_structured(soup) -> str:
    """HTML에서 구조화된 텍스트를 추출한다.

    - h1이 한글 이름이면 '성명: ' 프리픽스 추가
    - h2 섹션별로 콘텐츠를 구조화하여 AI 매핑 정확도 향상
    - 이력서, 회사소개서 등 문서형 HTML에 최적화
    """
    parts = []

    # 1. h1이 한글 이름(2-4자)이면 "성명: " 힌트 삽입
    h1_tag = soup.find("h1")
    if h1_tag:
        h1_text = h1_tag.get_text(strip=True)
        name_clean = re.sub(r"\s+", "", h1_text)
        if re.match(r"^[가-힣]{2,4}$", name_clean):
            parts.append(f"성명: {h1_text}")

    # 2. 헤더 영역의 "라벨: 값" 인라인 정보 수집 (연락처, 이메일 등)
    header_tag = soup.find("header")
    if header_tag:
        for elem in header_tag.find_all(["span", "p"]):
            if elem.find(["span", "p"]):  # 자식 있는 컨테이너 스킵
                continue
            text = elem.get_text(strip=True)
            if ":" in text and len(text) <= 80:
                key, _, val = text.partition(":")
                if key.strip() and val.strip() and len(key.strip()) <= 20 and "://" not in key:
                    parts.append(text)

    # 3. h2 섹션 구조가 있는 경우 섹션별 분리 추출
    sections = soup.find_all("section")
    if sections:
        for section in sections:
            h2 = section.find(["h2", "h3"])
            section_title = h2.get_text(strip=True) if h2 else ""
            if section_title:
                parts.append(f"\n[{section_title}]")

            # 반복 항목 감지: 클래스명이 "-item" 또는 "-entry"로 끝나는 최상위 div만
            # (career-header, career-title 등 하위 요소 제외)
            item_groups = section.find_all(
                lambda tag: tag.name in ("div", "article", "li") and
                any(c.endswith("-item") or c.endswith("-entry") or c in ("item", "entry")
                    for c in tag.get("class", []))
            )

            if item_groups:
                # 과거→현재 순 정렬: 기간 정보 파싱
                def _parse_start_year(item):
                    period_text = item.get_text()
                    m = re.search(r"(\d{4})\.\d{2}", period_text)
                    return int(m.group(1)) if m else 9999

                try:
                    item_groups_sorted = sorted(item_groups, key=_parse_start_year)
                except Exception:
                    item_groups_sorted = item_groups

                for idx, item in enumerate(item_groups_sorted, 1):
                    # 제목/기간/내용 요소 추출
                    title_elem = (
                        item.find(class_=re.compile(r"(title|name|position)", re.I))
                        or item.find(["h3", "h4"])  # 클래스 없는 h태그 폴백
                    )
                    period_elem = item.find(class_=re.compile(r"(period|date|duration|term)", re.I))
                    desc_elems = item.find_all(["li", "p"])

                    item_parts = [f"항목{idx})"]
                    if title_elem:
                        item_parts.append(f"직위/기관: {title_elem.get_text(strip=True)}")
                    if period_elem:
                        item_parts.append(f"기간: {period_elem.get_text(strip=True)}")
                    if desc_elems:
                        desc = " / ".join(
                            e.get_text(strip=True) for e in desc_elems
                            if e.get_text(strip=True)
                        )
                        if desc:
                            item_parts.append(f"내용: {desc}")
                    # 위 요소 모두 미발견 시 전체 텍스트 폴백
                    if len(item_parts) == 1:
                        item_parts.append(item.get_text(separator=" | ", strip=True))
                    parts.append(" | ".join(item_parts))
            else:
                # 반복 항목 없는 섹션: h4 기반 항목 분리 또는 전체 텍스트
                h4_items = section.find_all("h4")
                if h4_items:
                    for idx, h4 in enumerate(h4_items, 1):
                        item_text = h4.get_text(strip=True)
                        # h4 다음 형제 p/ul 콘텐츠 수집 (다음 h4 전까지)
                        content_parts = []
                        for sib in h4.next_siblings:
                            if getattr(sib, "name", None) == "h4":
                                break
                            if hasattr(sib, "get_text"):
                                t = sib.get_text(strip=True)
                                if t:
                                    content_parts.append(t)
                        content = " ".join(content_parts)[:300]
                        parts.append(f"항목{idx}) {item_text}" + (f" | 내용: {content}" if content else ""))
                else:
                    # 일반 섹션: 전체 텍스트 (h2/h3 제외)
                    if h2:
                        h2.extract()
                    parts.append(section.get_text(separator="\n", strip=True).strip())
    else:
        # 섹션 구조 없음: 전체 텍스트 그대로
        parts.append(soup.get_text(separator="\n", strip=True))

    return "\n".join(p for p in parts if p.strip())


def _read_content_file(file_path):
    """내용 파일을 읽어 텍스트로 반환한다."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext in (".txt", ".md", ".csv"):
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()

    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            lines = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    lines.append(" | ".join(cells))
            wb.close()
            return "\n".join(lines)
        except Exception as e:
            return f"[엑셀 읽기 실패: {e}]"

    if ext == ".docx":
        try:
            from docx import Document
            doc = Document(file_path)
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except Exception as e:
            return f"[워드 읽기 실패: {e}]"

    if ext in (".html", ".htm"):
        try:
            from bs4 import BeautifulSoup
            # charset 자동 감지: utf-8 시도 후 실패 시 euc-kr (카카오톡 내보내기 등)
            for enc in ("utf-8", "euc-kr", "cp949"):
                try:
                    with open(file_path, "r", encoding=enc) as f:
                        raw = f.read()
                    break
                except UnicodeDecodeError:
                    continue
            else:
                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                    raw = f.read()
            soup = BeautifulSoup(raw, "html.parser")
            for tag in soup(["script", "style"]):
                tag.decompose()

            return _extract_html_structured(soup)
        except Exception as e:
            return f"[HTML 읽기 실패: {e}]"

    if ext == ".json":
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return json.dumps(data, ensure_ascii=False, indent=2)

    # 기본: 텍스트로 읽기
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    except Exception:
        return f"[파일 읽기 실패: {file_path}]"


def _parse_kv_from_text(text: str) -> dict[str, str]:
    """텍스트에서 key:value 쌍을 파싱한다.

    지원 형식:
    - "이름: 홍길동"
    - "이름 홍길동" (공백 구분, 짧은 키)
    - 탭 구분 "이름\t홍길동"
    - 엑셀 행 "이름 | 홍길동 | ..."
    """
    kv = {}
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue

        # 콜론 구분 (가장 일반적)
        if ":" in line:
            key, _, val = line.partition(":")
            key, val = key.strip(), val.strip()
            # URL scheme("https", "http", "ftp" 등) 키 제외
            if key and val and len(key) <= 30 and "://" not in key:
                kv[key] = val
            continue

        # 탭 구분
        if "\t" in line:
            parts = [p.strip() for p in line.split("\t") if p.strip()]
            if len(parts) >= 2 and len(parts[0]) <= 30:
                kv[parts[0]] = parts[1]
            continue

        # 파이프 구분 (엑셀 변환 형식)
        if " | " in line:
            parts = [p.strip() for p in line.split(" | ") if p.strip()]
            if len(parts) >= 2 and len(parts[0]) <= 30:
                kv[parts[0]] = parts[1]
            continue

    return kv


def _parse_kv_from_excel(file_path: str) -> dict[str, str]:
    """엑셀 파일에서 key:value 구조를 직접 추출한다.

    지원 구조:
    - A열=키, B열=값 (2열 구조)
    - 헤더 행 자동 감지
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        kv = {}
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() if c is not None else "" for c in row]
                non_empty = [c for c in cells if c]
                if len(non_empty) >= 2:
                    key = non_empty[0]
                    val = non_empty[1]
                    if key and val and len(key) <= 30 and key not in ("None", ""):
                        kv[key] = val
        wb.close()
        return kv
    except Exception as e:
        print(f"[direct_map] 엑셀 직접 추출 실패: {e}")
        return {}


def direct_map(form_texts: list[str], content_paths: list[str], text: str = "") -> tuple[dict, str | None]:
    """AI 없이 문서에서 텍스트 추출 → 양식 필드명과 직접 매칭한다.

    Args:
        form_texts: 양식에서 추출된 텍스트 목록
        content_paths: 내용 파일 경로 목록
        text: 사용자가 입력한 텍스트 (선택)

    Returns:
        dict: {원본필드텍스트: 매핑된값} — 못 찾은 필드는 포함하지 않음
        str: 에러 메시지 (성공 시 None)
    """
    # 1. 모든 소스에서 key:value 수집
    all_kv: dict[str, str] = {}

    # 텍스트 입력
    if text and text.strip():
        all_kv.update(_parse_kv_from_text(text.strip()))

    # 파일별 처리
    for fp in content_paths:
        ext = os.path.splitext(fp)[1].lower()
        if ext in (".xlsx", ".xls"):
            # 엑셀: key:value 직접 추출 우선
            kv = _parse_kv_from_excel(fp)
            if kv:
                all_kv.update(kv)
            else:
                # 폴백: 텍스트로 변환 후 파싱
                raw = _read_content_file(fp)
                all_kv.update(_parse_kv_from_text(raw))
        else:
            raw = _read_content_file(fp)
            all_kv.update(_parse_kv_from_text(raw))

    if not all_kv and not text.strip():
        return None, "내용을 입력하거나 파일을 업로드해주세요."

    if not all_kv:
        return None, "파일에서 키:값 형식의 내용을 찾지 못했습니다. '항목명: 값' 형식으로 입력해주세요."

    print(f"[direct_map] key:value 파싱 완료: {len(all_kv)}개")

    # 2. 양식 필드와 매칭
    _SKIP = {"□", "☑", "※", "①", "②", "③", "④", "⑤", "⑥", "⑦", "⑧", "⑨", "⑩", "☐", "○", "●"}
    result: dict[str, str] = {}

    # 키 정규화 인덱스 (공백 제거 버전으로 매칭)
    kv_norm: dict[str, str] = {}
    for k, v in all_kv.items():
        kv_norm[k] = v
        kv_norm[re.sub(r"\s+", "", k)] = v

    # __N 접미사별 카운터 (같은 기본 필드 여러 개 순서대로 채우기)
    base_counter: dict[str, int] = {}

    for field in form_texts:
        if not field.strip() or field.strip() in _SKIP:
            continue

        # __N 접미사 처리
        base = re.sub(r"__\d+$", "", field)
        has_suffix = base != field

        if has_suffix:
            base_counter[base] = base_counter.get(base, 0) + 1
            # 소스에서 해당 순번 값 찾기
            # 먼저 "기본키__N" 직접 매칭 시도
            direct_key = field
            if direct_key in kv_norm:
                result[field] = kv_norm[direct_key]
                continue
            # 기본키로 첫 번째 값 사용 (순번=1일 때만)
            base_norm = re.sub(r"\s+", "", base)
            val = kv_norm.get(base) or kv_norm.get(base_norm)
            if val and base_counter[base] == 1:
                result[field] = val
            continue

        # 일반 필드: 정확 매칭
        field_norm = re.sub(r"\s+", "", field)
        val = kv_norm.get(field) or kv_norm.get(field_norm)
        if val:
            result[field] = val

    print(f"[direct_map] 매핑 결과: {len(result)}개 / {len(form_texts)}개 필드")

    if not result:
        return None, "양식 필드와 일치하는 항목을 찾지 못했습니다. 파일의 항목명이 양식과 일치하는지 확인해주세요."

    return result, None


def map_content(form_texts, user_content, content_file=None, structured=None,
                extra_content_files=None, extra_labels=None):
    """양식 필드와 사용자 내용을 AI로 매핑한다.

    Args:
        form_texts: 양식에서 추출된 텍스트 목록 (평면 리스트)
        user_content: 사용자가 입력한 텍스트
        content_file: 내용이 담긴 파일 경로 (선택)
        structured: extract_structured_fields()의 결과 (테이블 구조, 선택)
        extra_content_files: 추가 내용 파일 경로 리스트 (선택, 복수 파일)
        extra_labels: invoice_style 양식에서 일반 텍스트로 된 라벨을 [H]로
            취급하도록 추가 라벨 집합 전달 (InvoiceProcessor.INVOICE_LABELS)

    Returns:
        dict: {원본텍스트: 새텍스트} 또는 에러 시 None
        str: 에러 메시지 (성공 시 None)
    """
    api_key = _get_api_key()
    if not api_key:
        return None, "GEMINI_API_KEY가 설정되지 않았습니다. 환경변수 또는 .env 파일에 설정해주세요."

    # 내용 조합 (텍스트 + 파일 1개 + 추가 파일들)
    content_parts = []
    if user_content and user_content.strip():
        content_parts.append(user_content.strip())
    if content_file:
        file_text = _read_content_file(content_file)
        content_parts.append(file_text)
    if extra_content_files:
        for fp in extra_content_files:
            try:
                extra_text = _read_content_file(fp)
                content_parts.append(extra_text)
            except Exception as e:
                print(f"[ai/map] 추가 파일 읽기 실패: {e}")

    if not content_parts:
        return None, "내용을 입력하거나 파일을 업로드해주세요."

    combined_content = "\n\n".join(content_parts)
    print(f"[ai/map] 콘텐츠 합산: {len(content_parts)}개 소스, {len(combined_content):,}자")

    print(f"[ai/map] 통합 모드 (user_content len={len(user_content)})")

    # 구조화된 필드가 있으면 테이블 형식 프롬프트 사용
    use_structured = structured is not None and len(structured.get("tables", [])) > 0
    if use_structured:
        structured_text = _format_structured_fields(structured, extra_labels=extra_labels)
        print(f"[ai/map] 구조화 모드: 표 {len(structured['tables'])}개, 본문 {len(structured.get('paragraphs', []))}개")

    # 양식 필드 필터링 (평면 리스트 - 구조화 미사용 시 폴백)
    _SKIP = {"□", "☑", "※", "①", "②", "③", "④", "⑤", "⑥", "⑦", "⑧", "⑨", "⑩", "☐", "○", "●"}
    filtered_fields = [t for t in form_texts if 1 < len(t) <= 200 and t.strip() not in _SKIP]

    # 필드 상한: 3000개 (약 100장 분량)
    MAX_FIELDS = 3000
    if len(filtered_fields) > MAX_FIELDS:
        print(f"[ai/map] 필드 {len(filtered_fields)}개 → {MAX_FIELDS}개로 제한")
        filtered_fields = filtered_fields[:MAX_FIELDS]

    # 배치 분할 호출 (150개씩 나눠서 전체 커버)
    BATCH_SIZE = 150
    system_prompt = SYSTEM_PROMPT
    model_name = MODEL_NAME

    try:
        client = genai.Client(api_key=api_key)
        temperature = 0.1

        all_results = {}

        # 양식 필드를 배치로 분할
        if use_structured:
            struct_lines = structured_text.split("\n")
            field_batches = []
            current_batch = []
            current_lines = 0
            STRUCT_BATCH_LINES = 200
            for line in struct_lines:
                is_section_start = line.startswith("[표") or line.startswith("[본문")
                is_oversized = current_lines >= STRUCT_BATCH_LINES * 2
                if (is_section_start and current_lines >= STRUCT_BATCH_LINES) or is_oversized:
                    field_batches.append("\n".join(current_batch))
                    current_batch = []
                    current_lines = 0
                current_batch.append(line)
                current_lines += 1
            if current_batch:
                field_batches.append("\n".join(current_batch))
        else:
            field_batches = []
            for i in range(0, len(filtered_fields), BATCH_SIZE):
                batch = filtered_fields[i:i + BATCH_SIZE]
                field_batches.append("\n".join(f"- {t}" for t in batch))

        total_batches = len(field_batches)
        need_cache = total_batches > 1

        # 소형 문서 (1배치): 캐시 없이 1회 호출
        if not need_cache:
            prompt = USER_PROMPT.format(fields=field_batches[0], content=combined_content)
            print(f"[ai/map] 1회 호출 (캐시 불필요, 필드 {len(filtered_fields)}개, use_structured={use_structured})")
            try:
                print(f"[ai/map] fields 앞부분:\n{field_batches[0][:500]}")
                print(f"[ai/map] content 앞부분:\n{combined_content[:500]}")
            except UnicodeEncodeError:
                pass  # Windows cp949 콘솔 출력 실패는 무시 (특수문자 포함 시 발생)

            response = _call_with_retry(client, model_name, prompt, system_prompt, temperature)

            try:
                print(f"[ai/map] AI 응답 앞부분:\n{response.text[:800] if response.text else '(빈 응답)'}")
            except UnicodeEncodeError:
                pass
            parsed = _parse_json_response(response.text)
            if parsed:
                all_results = _collect_results(parsed)
                print(f"[ai/map] 파싱 결과: {len(all_results)}개 키")

        # 대형 문서 (다배치): Explicit Caching 사용
        else:
            cache = None
            try:
                # 캐시 생성: 시스템 프롬프트 + 콘텐츠 (TTL 10분)
                cache = client.caches.create(
                    model=model_name,
                    config=types.CreateCachedContentConfig(
                        system_instruction=system_prompt,
                        contents=[types.Content(role="user", parts=[types.Part(text=combined_content)])],
                        ttl="600s",
                    )
                )
                print(f"[ai/map] 캐시 생성 완료: {total_batches}배치, TTL=10분")

                for batch_idx, batch_text in enumerate(field_batches):
                    prompt_template = USER_PROMPT
                    # 캐시에 콘텐츠가 있으므로 프롬프트에서는 양식만
                    prompt = prompt_template.format(fields=batch_text, content="(위에 제공된 내용 참조)")

                    print(f"[ai/map] cached batch {batch_idx+1}/{total_batches}")

                    if batch_idx > 0:
                        import time
                        time.sleep(0.5)

                    # 캐시 유효성 체크 + 재생성
                    try:
                        client.caches.get(name=cache.name)
                    except Exception:
                        print("[ai/map] 캐시 만료, 재생성")
                        cache = client.caches.create(
                            model=model_name,
                            config=types.CreateCachedContentConfig(
                                system_instruction=system_prompt,
                                contents=[types.Content(role="user", parts=[types.Part(text=combined_content)])],
                                ttl="600s",
                            )
                        )

                    response = _call_cached_with_retry(client, model_name, cache.name, prompt, temperature)
                    parsed = _parse_json_response(response.text)
                    if parsed:
                        for k, v in _collect_results(parsed).items():
                            all_results[k] = v

            except Exception as cache_err:
                if "cache" in str(cache_err).lower() or "CachedContent" in str(cache_err):
                    # 캐시 기능 실패 → 폴백: 캐시 없이 반복 전송
                    print(f"[ai/map] 캐시 실패, 폴백 모드: {cache_err}")
                    for batch_idx, batch_text in enumerate(field_batches):
                        prompt_template = USER_PROMPT
                        prompt = prompt_template.format(fields=batch_text, content=combined_content)
                        print(f"[ai/map] fallback batch {batch_idx+1}/{total_batches}")
                        if batch_idx > 0:
                            import time
                            time.sleep(0.5)
                        response = _call_with_retry(client, model_name, prompt, system_prompt, temperature)
                        parsed = _parse_json_response(response.text)
                        if parsed:
                            for k, v in _collect_results(parsed).items():
                                all_results[k] = v
                else:
                    raise
            finally:
                # 3중 방어 (1): 즉시 삭제
                if cache:
                    try:
                        client.caches.delete(name=cache.name)
                        print("[ai/map] 캐시 삭제 완료")
                    except Exception:
                        pass  # (2): TTL 10분 자동 만료가 백업

        if not all_results:
            return None, "매핑할 항목을 찾지 못했습니다. 내용을 더 구체적으로 입력해주세요."

        # [Fix 1+5] AI 환각 키 정규화
        field_set = set(form_texts)
        # 공백/이스케이프 정규화된 역방향 인덱스
        _norm_index = {}
        for t in form_texts:
            _norm_index[t] = t
            _norm_index[t.replace("&lt;", "<").replace("&gt;", ">")] = t
            _norm_index[t.replace("<", "&lt;").replace(">", "&gt;")] = t
            _norm_index[re.sub(r"\s+", "", t)] = t  # 공백 제거 버전

        normalized = {}
        for k, v in all_results.items():
            if k in field_set:
                normalized[k] = v
                continue

            # 1단계: 이스케이프 변환
            matched = _norm_index.get(k) or _norm_index.get(
                k.replace("<", "&lt;").replace(">", "&gt;")
            ) or _norm_index.get(
                k.replace("&lt;", "<").replace("&gt;", ">")
            )

            # 2단계: 공백 제거 매칭
            if not matched:
                k_no_space = re.sub(r"\s+", "", k)
                matched = _norm_index.get(k_no_space)

            # 3단계: 접두사 매칭 (AI가 "- " 같은 접두사를 빼먹은 경우)
            if not matched:
                for t in form_texts:
                    if t.endswith(k) or k.endswith(t):
                        if min(len(k), len(t)) / max(len(k), len(t)) > 0.8:
                            matched = t
                            break

            normalized[matched or k] = v

        # [Fix 3] 미매핑 필드 재시도: 1차에서 누락된 필드만 모아서 2차 호출
        # 구조화 모드에서는 재시도 스킵 (평면 형식으로 보내면 라벨 보호 무효화됨)
        mapped_keys = set(normalized.keys())
        unmapped = [f for f in filtered_fields if f not in mapped_keys]
        if not use_structured and unmapped and len(unmapped) > 5:
            retry_batches = (len(unmapped) + BATCH_SIZE - 1) // BATCH_SIZE
            print(f"[ai/map] 미매핑 {len(unmapped)}개 재시도 ({retry_batches} batches)")
            for rb in range(min(retry_batches, 3)):  # 최대 3배치만 재시도
                start = rb * BATCH_SIZE
                batch = unmapped[start:start + BATCH_SIZE]
                fields_text = "\n".join(f"- {t}" for t in batch)
                retry_prompt = USER_PROMPT.format(fields=fields_text, content=combined_content)

                import time
                time.sleep(1)
                try:
                    retry_resp = _call_with_retry(
                        client, model_name, retry_prompt, system_prompt, temperature
                    )
                    retry_parsed = _parse_json_response(retry_resp.text)
                    if retry_parsed:
                        for k2, v2 in _collect_results(retry_parsed).items():
                            matched2 = _norm_index.get(k2) or _norm_index.get(
                                k2.replace("<", "&lt;").replace(">", "&gt;")
                            ) or _norm_index.get(
                                k2.replace("&lt;", "<").replace("&gt;", ">")
                            )
                            normalized[matched2 or k2] = v2
                    print(f"[ai/map] 재시도 batch {rb+1}: +{len(retry_parsed or {})}개")
                except Exception as retry_e:
                    print(f"[ai/map] 재시도 실패: {retry_e}")

        # ── 한국 포맷터 후처리 (금액 콤마, 날짜 년월일, 전화 하이픈, 사업자번호) ──
        try:
            from kr_formatter import KrFormatter
            fmt_result = KrFormatter.auto_detect_and_format(normalized)
            normalized = fmt_result["formatted"]
            for log_entry in fmt_result["log"]:
                try:
                    print(f"[ai/map] {log_entry}")
                except UnicodeEncodeError:
                    pass
        except ImportError:
            print("[ai/map] kr_formatter 미설치, 포맷팅 스킵")
        except Exception as fmt_err:
            print(f"[ai/map] 포맷팅 실패 (원본 유지): {fmt_err}")

        # ── 정밀 필드 소스 대조 검증 (할루시네이션 방지) ──
        _verify_exact_fields(normalized, combined_content)

        return normalized, None

    except Exception as e:
        error_msg = str(e)
        if "API_KEY" in error_msg or "401" in error_msg:
            return None, "AI 서비스 연결에 문제가 있습니다. 관리자에게 문의해주세요."
        if "429" in error_msg:
            return None, "요청이 많아 잠시 처리가 지연됩니다. 1분 후 다시 시도해주세요."
        if "timeout" in error_msg.lower() or "deadline" in error_msg.lower():
            return None, "AI 처리 시간이 초과되었습니다. 다시 시도해주세요."
        return None, "AI 처리 중 문제가 발생했습니다. 다시 시도해주세요."
