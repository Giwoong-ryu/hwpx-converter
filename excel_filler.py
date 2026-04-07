"""엑셀 양식 채우기 - 엑셀 양식 + 데이터 엑셀 → 완성 엑셀 N개"""

import os
import shutil
import tempfile
import zipfile

try:
    import openpyxl
    from copy import copy
except ImportError:
    openpyxl = None


def extract_excel_texts(xlsx_path):
    """엑셀에서 텍스트가 있는 셀 목록을 추출한다.

    Returns: list of {"sheet": str, "cell": str, "value": str}
    """
    if not openpyxl:
        return []

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    texts = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip():
                    texts.append({
                        "sheet": ws.title,
                        "cell": cell.coordinate,
                        "value": str(cell.value).strip(),
                    })
    wb.close()
    return texts


def fill_excel(template_path, replacements, output_path=None):
    """엑셀 양식의 텍스트를 치환하여 새 파일을 생성한다.

    Args:
        template_path: 양식 엑셀 경로
        replacements: {원본텍스트: 새텍스트} 딕셔너리
        output_path: 출력 경로 (None이면 자동 생성)

    Returns: output_path
    """
    if not openpyxl:
        raise RuntimeError("openpyxl이 설치되지 않았습니다.")

    if not output_path:
        output_path = os.path.join(tempfile.mkdtemp(), "filled.xlsx")

    # 원본 복사
    shutil.copy2(template_path, output_path)

    # 열어서 치환 (수식 셀은 건드리지 않음)
    wb = openpyxl.load_workbook(output_path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                # 수식 셀(= 로 시작) 보호: 텍스트 치환 대상에서 제외
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    continue
                val = str(cell.value)
                for old, new in replacements.items():
                    if old in val:
                        val = val.replace(old, new)
                if val != str(cell.value):
                    cell.value = val
    wb.save(output_path)
    wb.close()
    return output_path


def batch_fill_excel(template_path, data_path):
    """엑셀 양식 + 데이터 엑셀 → N개 엑셀 생성.

    데이터 엑셀 형식:
      1행: 양식에 있는 바꿀 원본 텍스트
      2행~: 각 문서에 넣을 새 텍스트

    Returns: (zip_path, count, error)
    """
    if not openpyxl:
        return None, 0, "openpyxl이 설치되지 않았습니다."

    wb = openpyxl.load_workbook(data_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return None, 0, "데이터 엑셀에 최소 2행(헤더 + 데이터)이 필요합니다."

    headers = [str(h).strip() if h else "" for h in rows[0]]
    if not any(headers):
        return None, 0, "1행(헤더)이 비어있습니다."

    out_dir = tempfile.mkdtemp()
    generated = []
    used_names: set = set()

    for i, row in enumerate(rows[1:], start=1):
        replacements = {}
        for j, val in enumerate(row):
            if j < len(headers) and headers[j] and val is not None:
                replacements[headers[j]] = str(val).strip()
        if not replacements:
            continue

        first_val = str(row[0]).strip() if row[0] else f"문서_{i}"
        safe_name = "".join(c for c in first_val if c not in r'\/:*?"<>|')[:50] or f"문서_{i}"
        candidate = f"{safe_name}.xlsx"
        if candidate in used_names:
            candidate = f"{safe_name}_{i}.xlsx"
        used_names.add(candidate)
        out_path = os.path.join(out_dir, candidate)

        fill_excel(template_path, replacements, out_path)
        generated.append(out_path)

    if not generated:
        return None, 0, "생성할 데이터가 없습니다."

    zip_path = os.path.join(tempfile.mkdtemp(), "DocFlow_excel_batch.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for fp in generated:
            zf.write(fp, os.path.basename(fp))

    shutil.rmtree(out_dir, ignore_errors=True)
    return zip_path, len(generated), None
