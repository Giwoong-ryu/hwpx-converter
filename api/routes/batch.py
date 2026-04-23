"""대량 생성 API - 엑셀 헤더 AI 매핑 + 대량 생성"""

import json
import os
import shutil
import tempfile
import zipfile

from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from fastapi.responses import FileResponse
from pydantic import BaseModel

from api.services.file_manager import file_manager
from api.services.metrics import log as mlog, Timer
from clone_form import extract_texts, clone as clone_hwpx

router = APIRouter()


@router.post("/map-headers")
async def map_excel_headers(
    file_id: str = Form(...),
    excel: UploadFile = File(...),
):
    """엑셀 헤더와 양식 텍스트를 AI로 매핑"""
    form_path = file_manager.get_path(file_id)
    if not form_path:
        raise HTTPException(status_code=404, detail="양식 파일을 찾을 수 없습니다.")

    eid = await file_manager.save_upload(excel)
    excel_path = file_manager.get_path(eid)

    # 엑셀 헤더 읽기
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="엑셀에 최소 2행(헤더 + 데이터)이 필요합니다.")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    sample_row = [str(v).strip() if v else "" for v in rows[1]]
    row_count = len(rows) - 1

    # HWP 파일은 extract_texts 미지원 (HWPX만 가능)
    if form_path.lower().endswith(".hwp") and not form_path.lower().endswith(".hwpx"):
        raise HTTPException(status_code=400, detail="대량 생성은 HWPX 파일만 지원합니다. HWP 파일을 HWPX로 변환 후 사용해주세요.")

    # 양식 텍스트
    form_texts = extract_texts(form_path)

    # AI로 헤더 ↔ 양식 텍스트 매핑
    from ai_mapper import _get_api_key, _parse_json_response
    from google import genai
    from google.genai import types

    api_key = _get_api_key()
    if not api_key:
        raise HTTPException(status_code=500, detail="AI 서비스 설정에 문제가 있습니다. 관리자에게 문의해주세요.")

    prompt = f"""양식 문서의 텍스트 목록과 엑셀의 헤더+샘플 데이터가 있습니다.
각 엑셀 헤더가 양식의 어떤 텍스트에 해당하는지 매핑해주세요.

양식 텍스트 (일부):
{json.dumps(form_texts[:200], ensure_ascii=False)}

엑셀 헤더: {json.dumps(headers, ensure_ascii=False)}
엑셀 샘플 (1행): {json.dumps(sample_row, ensure_ascii=False)}

JSON 배열로 응답: [{{"header": "엑셀헤더", "form_text": "양식에서 매칭되는 텍스트"}}]
매칭 안 되는 헤더는 form_text를 빈 문자열로.
"""
    try:
        client = genai.Client(api_key=api_key)
        # 대량 단순 매핑용 모델 (기본: 2.5-flash, env로 2.5-flash-lite 등 교체 가능)
        import os as _os
        _batch_model = _os.getenv("DOCFLOW_BATCH_MODEL", "gemini-2.5-flash")
        response = client.models.generate_content(
            model=_batch_model,
            contents=prompt,
            config=types.GenerateContentConfig(temperature=0.1),
        )
        parsed = _parse_json_response(response.text)
        if isinstance(parsed, list):
            mappings = parsed
        elif isinstance(parsed, dict):
            mappings = parsed.get("mappings", parsed.get("매핑", []))
        else:
            mappings = []
    except Exception as e:
        raise HTTPException(status_code=500, detail="AI 매핑 처리 중 문제가 발생했습니다. 잠시 후 다시 시도해주세요.")

    return {
        "excel_id": eid,
        "headers": headers,
        "sample_row": sample_row,
        "row_count": row_count,
        "mappings": mappings,
    }


class BatchGenerateRequest(BaseModel):
    file_id: str
    excel_id: str
    column_mappings: list  # [{"header": "이름", "form_text": "홍길동"}, ...]


@router.post("/generate-mapped")
def batch_generate_mapped(req: BatchGenerateRequest):
    """AI 매핑 결과를 기반으로 대량 생성"""
    form_path = file_manager.get_path(req.file_id)
    if not form_path:
        raise HTTPException(status_code=404, detail="양식 파일을 찾을 수 없습니다.")

    # HWP 파일은 clone_hwpx 미지원 (HWPX만 가능)
    if form_path.lower().endswith(".hwp") and not form_path.lower().endswith(".hwpx"):
        raise HTTPException(status_code=400, detail="대량 생성은 HWPX 파일만 지원합니다. HWP 파일을 HWPX로 변환 후 사용해주세요.")

    excel_path = file_manager.get_path(req.excel_id)
    if not excel_path:
        raise HTTPException(status_code=404, detail="엑셀 파일을 찾을 수 없습니다.")

    # 매핑 테이블: header_index -> form_text
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="엑셀에 최소 2행(헤더 + 데이터)이 필요합니다.")
    headers = [str(h).strip() if h else "" for h in rows[0]]

    header_to_form = {}
    for m in req.column_mappings:
        h = m.get("header", "")
        ft = m.get("form_text", "")
        if h and ft:
            header_to_form[h] = ft

    out_dir = tempfile.mkdtemp()
    zip_dir = tempfile.mkdtemp()
    zip_path = os.path.join(zip_dir, "DocFlow_batch.zip")  # 사전 정의 (clone_hwpx 예외 시 UnboundLocalError 방지)
    generated = []
    used_names: set[str] = set()

    try:
        for i, row in enumerate(rows[1:], start=1):
            replacements = {}
            for j, val in enumerate(row):
                if j < len(headers) and headers[j] in header_to_form and val is not None:
                    form_text = header_to_form[headers[j]]
                    replacements[form_text] = str(val).strip()

            if not replacements:
                continue

            first_val = str(row[0]).strip() if row[0] else f"문서_{i}"
            safe_name = "".join(c for c in first_val if c not in r'\/:*?"<>|')[:50] or f"문서_{i}"

            # 파일명 중복 처리: 겹치면 _i 접미사 붙여 고유하게
            candidate = f"{safe_name}.hwpx"
            if candidate in used_names:
                candidate = f"{safe_name}_{i}.hwpx"
            used_names.add(candidate)
            out_path = os.path.join(out_dir, candidate)

            clone_hwpx(form_path, out_path, replacements=replacements)
            generated.append(out_path)

        if not generated:
            mlog("batch", success=False, error="생성할 데이터 없음")
            raise HTTPException(status_code=400, detail="생성할 데이터가 없습니다.")

        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for fp in generated:
                zf.write(fp, os.path.basename(fp))

    finally:
        shutil.rmtree(out_dir, ignore_errors=True)

    mlog("batch", success=True, field_count=len(generated), detail=f"docs={len(generated)}")
    return FileResponse(
        zip_path,
        filename="DocFlow_batch.zip",
        media_type="application/zip",
        background=None,  # FileResponse가 반환된 후 zip_dir는 OS가 정리
    )


# 기존 API도 유지 (하위 호환)
@router.post("/generate")
async def batch_gen(
    file_id: str = Form(...),
    excel: UploadFile = File(...),
):
    from features import batch_generate
    path = file_manager.get_path(file_id)
    if not path:
        raise HTTPException(status_code=404, detail="양식 파일을 찾을 수 없습니다.")
    eid = await file_manager.save_upload(excel)
    excel_path = file_manager.get_path(eid)
    with Timer() as t:
        zp, cnt, err = batch_generate(path, excel_path)
    if err:
        mlog("batch_legacy", success=False, duration_ms=t.ms, error=err)
        raise HTTPException(status_code=400, detail=err)
    mlog("batch_legacy", success=True, field_count=cnt, duration_ms=t.ms)
    return FileResponse(zp, filename="DocFlow_batch.zip", media_type="application/zip")
