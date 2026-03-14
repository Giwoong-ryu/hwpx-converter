"""
Excel/CSV -> HWPX 구조 JSON 변환
openpyxl로 병합 셀 정보까지 파싱
"""
import csv
import json
import os
from pathlib import Path


def parse_excel(file_path: str) -> dict:
    """Excel 파일 -> HWPX 구조 JSON"""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    merged_ranges = list(ws.merged_cells.ranges)
    merge_map = {}
    for mr in merged_ranges:
        min_row, min_col = mr.min_row, mr.min_col
        max_row, max_col = mr.max_row, mr.max_col
        rowspan = max_row - min_row + 1
        colspan = max_col - min_col + 1
        merge_map[(min_row, min_col)] = (rowspan, colspan)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if (r, c) != (min_row, min_col):
                    merge_map[(r, c)] = "skip"

    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row,
                              min_col=1, max_col=ws.max_column))
    if not rows:
        return {"document": {"title": "Empty", "page_width_hu": 42520}, "sections": []}

    n_rows = len(rows)
    n_cols = len(rows[0])

    col_ratios = [round(100 / n_cols, 1)] * n_cols
    remainder = 100 - sum(col_ratios)
    col_ratios[-1] = round(col_ratios[-1] + remainder, 1)

    cells = []
    title = ""
    for ri, row in enumerate(rows):
        for ci, cell in enumerate(row):
            excel_pos = (ri + 1, ci + 1)

            if merge_map.get(excel_pos) == "skip":
                continue

            text = str(cell.value) if cell.value is not None else ""
            if text == "None":
                text = ""

            style = _extract_cell_style(cell, ri)

            cell_data = {
                "row": ri,
                "col": ci,
                "text": text,
                "style": style,
            }

            merge_info = merge_map.get(excel_pos)
            if merge_info and merge_info != "skip":
                rowspan, colspan = merge_info
                if rowspan > 1:
                    cell_data["rowspan"] = rowspan
                if colspan > 1:
                    cell_data["colspan"] = colspan

            if ri == 0 and text and len(text) > len(title):
                title = text

            cells.append(cell_data)

    table_section = {
        "type": "table",
        "table": {
            "rows": n_rows,
            "cols": n_cols,
            "col_widths_ratio": col_ratios,
            "cells": cells,
        }
    }

    return {
        "document": {"title": title, "page_width_hu": 42520},
        "sections": [table_section],
    }


def _extract_cell_style(cell, row_idx: int) -> dict:
    """openpyxl 셀 -> 스타일 dict"""
    style = {}

    if cell.font:
        if cell.font.bold:
            style["bold"] = True
        if cell.font.size and cell.font.size != 11:
            style["font_size_pt"] = int(cell.font.size)
        if cell.font.color and cell.font.color.rgb and cell.font.color.rgb != "00000000":
            rgb = str(cell.font.color.rgb)
            if len(rgb) == 8:
                rgb = rgb[2:]
            style["text_color"] = f"#{rgb}"

    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
        rgb = str(cell.fill.fgColor.rgb)
        if rgb not in ("00000000", "FFFFFFFF", "00FFFFFF") and len(rgb) >= 6:
            if len(rgb) == 8:
                rgb = rgb[2:]
            style["bg_color"] = f"#{rgb}"

    align = "LEFT"
    if cell.alignment and cell.alignment.horizontal:
        align_map = {"center": "CENTER", "right": "RIGHT", "left": "LEFT"}
        align = align_map.get(cell.alignment.horizontal, "LEFT")
    elif row_idx == 0:
        align = "CENTER"
    style["align"] = align

    if row_idx == 0:
        style["is_header"] = True
        if "bold" not in style:
            style["bold"] = True

    return style


def parse_csv(file_path: str) -> dict:
    """CSV 파일 -> HWPX 구조 JSON"""
    with open(file_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        all_rows = list(reader)

    if not all_rows:
        return {"document": {"title": "Empty", "page_width_hu": 42520}, "sections": []}

    n_rows = len(all_rows)
    n_cols = max(len(row) for row in all_rows)

    col_ratios = [round(100 / n_cols, 1)] * n_cols
    remainder = 100 - sum(col_ratios)
    col_ratios[-1] = round(col_ratios[-1] + remainder, 1)

    cells = []
    title = ""
    for ri, row in enumerate(all_rows):
        for ci in range(n_cols):
            text = row[ci] if ci < len(row) else ""
            style = {"align": "CENTER" if ri == 0 else "LEFT"}
            if ri == 0:
                style["bold"] = True
                style["is_header"] = True
                if text and len(text) > len(title):
                    title = text

            cells.append({
                "row": ri,
                "col": ci,
                "text": text,
                "style": style,
            })

    table_section = {
        "type": "table",
        "table": {
            "rows": n_rows,
            "cols": n_cols,
            "col_widths_ratio": col_ratios,
            "cells": cells,
        }
    }

    return {
        "document": {"title": title, "page_width_hu": 42520},
        "sections": [table_section],
    }


def parse_file(file_path: str) -> dict:
    """파일 확장자에 따라 적절한 파서 호출"""
    ext = Path(file_path).suffix.lower()
    if ext in (".xlsx", ".xls"):
        return parse_excel(file_path)
    elif ext == ".csv":
        return parse_csv(file_path)
    else:
        raise ValueError(f"지원하지 않는 파일 형식: {ext}")
